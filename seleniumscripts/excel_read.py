import openpyxl
# excel_workbook=openpyxl.load_workbook("C:\\Users\\rashmi\\python_selenium_advanced\\seleniumscripts\\sample.xlsx")
# #excel_workbook.create_sheet("sheet1",0)
# sheet=excel_workbook.get_sheet_by_name("Sheet1")
# print(sheet['A2'].value)
# value=sheet.cell(row=1,column=1).value
# print(value)

#write into excel
excel_workbook = openpyxl.load_workbook("C:\\Users\\rashmi\\python_selenium_advanced\\seleniumscripts\\sample.xlsx")
sheet=excel_workbook.get_sheet_by_name("Sheet1")
w=sheet['A3'].value=20
excel_workbook.save("C:\\Users\\rashmi\\python_selenium_advanced\\seleniumscripts\\sample.xlsx")
print(w)