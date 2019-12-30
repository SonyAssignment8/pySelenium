# Read File :
import openpyxl
excel_workbook = openpyxl.load_workbook("D:\\New folder\\Book1.xlsx")
# excel_workbook1 = openpyxl.load_workbook("D:\\New folder\\Book2.xlsx")
excel_workbook.create_sheet("Book1.xlsx",0)
sheet = excel_workbook.get_sheet_by_name("Sheet1")
# sheet1 = excel_workbook.get_sheet_by_name("Sheet1")
x = sheet.cell(row=1, column=1).value
y = sheet.cell(row=1, column=2).value
p = sheet.cell(row=2, column=1).value
q = sheet.cell(row=2, column=2).value
# s = sheet.cell(row=1, column=1).value
print(x)
print(y)
print(p)
print(q)
# print(s)

