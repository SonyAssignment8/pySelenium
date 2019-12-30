# Wrire in Excel

import openpyxl
excel_workbook = openpyxl.load_workbook("D:\\New folder\\Book1.xlsx")
# excel_workbook1 = openpyxl.load_workbook("D:\\New folder\\Book2.xlsx")
excel_workbook.create_sheet("Book1.xlsx",0)
sheet = excel_workbook.get_sheet_by_name("Sheet1")
# sheet1 = excel_workbook.get_sheet_by_name("Sheet1")
# x = sheet.cell(row=4, column=2).value = 20
# print(x)
# or we can use
sheet['B4'].value = 20
print(sheet['B4'].value)
excel_workbook.save("D:\\New folder\\Book1.xlsx")
