import openpyxl

# To read data from excel(.xlsx)
excel_workbook = openpyxl.load_workbook("D:\\TestCase_Repository.xlsx")
#excel_workbook.create_sheet("sheet1", 0)
sheet = excel_workbook.get_sheet_by_name("Allocation")
print(sheet['A2'].value)
a = sheet.cell(row=1, column=1).value

# to write data from excel(.xlsx)
file_name = 'TestCase_Repository.xlsx'
excel_workbook = openpyxl.load_workbook(file_name)
#excel_workbook.create_sheet("sheet1", 0)
sheet = excel_workbook.get_sheet_by_name("Sheet1")
sheet['A2'].value = 20
sheet.cell(row=1, column=1).value = 40
excel_workbook.save(file_name)
print(sheet['A2'].value)
print(sheet.cell(row=1, column=1).value)
